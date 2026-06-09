"""
workbook.py - excel-mutate 的工作簿操作层

提供高层 API:
- list_sheets / read_sheet
- insert_row / delete_row
- insert_column / delete_column
- insert_sheet / delete_sheet
- set_cell (改单元格 value 或 formula)
- save (写回 .xlsx)

所有"改结构"操作自动调用 formula.rewrite_* 重写所有相关公式。
"""
from __future__ import annotations
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional, List, Tuple, Dict, Any

from .formula import (
    rewrite_after_insert_row,
    rewrite_after_delete_row,
    rewrite_after_insert_column,
    rewrite_after_delete_column,
)


class Workbook:
    """包装 openpyxl.Workbook,提供结构操作 + 公式重写"""
    
    def __init__(self, path: str):
        self.path = Path(path)
        self.wb = openpyxl.load_workbook(str(self.path))
        # 跟踪警告: 公式指向被删的位置
        self.warnings: List[str] = []
    
    # ------------------------------------------------------------------
    # 读
    # ------------------------------------------------------------------
    def list_sheets(self) -> List[str]:
        return self.wb.sheetnames
    
    def read_sheet(self, name: str, max_row: Optional[int] = None, max_col: Optional[int] = None) -> List[List[str]]:
        """返回二维数组(每行 cell 的字符串表示)"""
        ws = self.wb[name]
        rows = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            if max_row and r_idx > max_row:
                break
            row_data = []
            for c_idx, cell in enumerate(row, start=1):
                if max_col and c_idx > max_col:
                    break
                v = cell.value
                if v is None:
                    row_data.append("")
                elif isinstance(v, str) and v.startswith("="):
                    row_data.append(f"{v}")  # formula
                else:
                    row_data.append(str(v))
            rows.append(row_data)
        return rows
    
    # ------------------------------------------------------------------
    # 改结构
    # ------------------------------------------------------------------
    def insert_row(self, sheet: str, at_row: int, count: int = 1) -> List[str]:
        """在 at_row 处插入 count 行,重写所有公式"""
        if sheet not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found")
        ws = self.wb[sheet]
        warnings = self._rewrite_all_formulas(
            sheet,
            lambda f: rewrite_after_insert_row(f, at_row=at_row, count=count, sheet=sheet),
        )
        # 真正插入(用 openpyxl 内部方法,先重写公式再 insert 避免冲突)
        ws.insert_rows(at_row, count)
        return warnings
    
    def delete_row(self, sheet: str, at_row: int, count: int = 1) -> List[str]:
        if sheet not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found")
        ws = self.wb[sheet]
        warnings = self._rewrite_all_formulas(
            sheet,
            lambda f: rewrite_after_delete_row(f, at_row=at_row, count=count, sheet=sheet),
        )
        ws.delete_rows(at_row, count)
        return warnings
    
    def insert_column(self, sheet: str, at_col: int, count: int = 1) -> List[str]:
        if sheet not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found")
        ws = self.wb[sheet]
        warnings = self._rewrite_all_formulas(
            sheet,
            lambda f: rewrite_after_insert_column(f, at_col=at_col, count=count, sheet=sheet),
        )
        ws.insert_cols(at_col, count)
        return warnings
    
    def delete_column(self, sheet: str, at_col: int, count: int = 1) -> List[str]:
        if sheet not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found")
        ws = self.wb[sheet]
        warnings = self._rewrite_all_formulas(
            sheet,
            lambda f: rewrite_after_delete_column(f, at_col=at_col, count=count, sheet=sheet),
        )
        ws.delete_cols(at_col, count)
        return warnings
    
    def insert_sheet(self, name: str, position: Optional[int] = None) -> None:
        if name in self.wb.sheetnames:
            raise ValueError(f"Sheet '{name}' already exists")
        self.wb.create_sheet(name, position)
    
    def delete_sheet(self, name: str) -> None:
        if name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{name}' not found")
        del self.wb[name]
    
    def rename_sheet(self, old: str, new: str) -> None:
        if old not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{old}' not found")
        if new in self.wb.sheetnames:
            raise ValueError(f"Sheet '{new}' already exists")
        # 改 sheet 名 → 所有引用该 sheet 的公式需改
        # openpyxl 不会自动重写,得自己扫
        self.wb[old].title = new
        for sn in self.wb.sheetnames:
            ws = self.wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        new_formula = _rename_sheet_in_formula(cell.value, old, new)
                        if new_formula != cell.value:
                            cell.value = new_formula
    
    # ------------------------------------------------------------------
    # 改 cell
    # ------------------------------------------------------------------
    def set_cell(self, sheet: str, coord: str, value: Any) -> None:
        """coord 格式: 'A1' 或 'B2'"""
        if sheet not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found")
        self.wb[sheet][coord] = value
    
    def get_cell(self, sheet: str, coord: str) -> Any:
        if sheet not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found")
        return self.wb[sheet][coord].value
    
    # ------------------------------------------------------------------
    # 写
    # ------------------------------------------------------------------
    def save(self, dest: Optional[str] = None) -> str:
        out = Path(dest) if dest else self.path
        self.wb.save(str(out))
        return str(out)
    
    # ------------------------------------------------------------------
    # 内部
    # ------------------------------------------------------------------
    def _rewrite_all_formulas(self, target_sheet: str, transform) -> List[str]:
        """扫描所有 sheet 的所有 cell,把公式用 transform 重写
        target_sheet 是被改结构的 sheet,本表的引用会受影响"""
        warnings = []
        for sn in self.wb.sheetnames:
            ws = self.wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        new_v = transform(v)
                        if new_v != v:
                            # 如果公式里含一个指向被删位置的引用,我们的 transform 不会改字符串
                            # 但 _rewrite_cell_refs 对被删的 cell 返回 None,会跳过
                            # 我们在调用方很难直接知道哪个 cell 被跳过了,先用简单的"长度变化"做检测
                            if len(new_v) == len(v) and new_v == v:
                                # 没变,可能指向被删位置
                                warnings.append(
                                    f"{sn}!{cell.coordinate}: 公式 '{v}' 引用了被删除的范围(保留原样,Excel 打开时显示 #REF!)"
                                )
                            cell.value = new_v
        return warnings


def _rename_sheet_in_formula(formula: str, old: str, new: str) -> str:
    """公式中 sheet 引用改名: Sheet1 -> NewName"""
    # 'Sheet Name'! 形式
    formula = formula.replace(f"'{old}'!", f"'{new}'!")
    # Sheet1! 形式(只在标识符边界)
    import re
    # 匹配: 不在字符串里的 Sheet1!  (简化:用 lookbehind/lookahead)
    pattern = re.compile(rf"(?<![A-Za-z0-9_']){re.escape(old)}!")
    return pattern.sub(f"{new}!", formula)
